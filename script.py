# Create a script that opens a text document and reads all lines of the document.
# Use the re.findall() function to find all the 10 digit numbers in the document.
# Print the numbers in the order in which they appear in the document.
# Print the total number of 10 digit numbers in the document.
# Create a method for each specific operation
# The main method should be a loop through all the txt documents in the directory until it gets to the last one.
# Each loop should write the list to an excel file, each list should have it's own column.


import re
import os
import openpyxl


def read_file(file_name):
    """
    This function reads the file and returns a list of all the lines in the file.
    :param file_name:
    :return:
    """
    with open(file_name, 'r') as f:
        return f.readlines()


def find_numbers(lines):
    """
    This function finds all the numbers in the lines of the file.
    :param lines:
    :return:
    """
    numbers = []
    for line in lines:
        numbers.extend(re.findall('\d{10}', line))
    return numbers

# Create a method that creates a new excel file called doc.xls
# And call the xls file 'doc.xls'
def create_excel_file():
    # Check to see if doc.xlsx exists. If it does, don't create a new one.
    # If it doesn't, create a new one.
    if os.path.exists('doc.xlsx'):
        # If it exists, delete the file
        os.remove('doc.xlsx')
        # Create the excel file with a sheet called 'Main'
        book = openpyxl.Workbook()
        book.save('doc.xlsx')
        book.close()
    else:
        # Create the excel file with a sheet called 'Main'
        book = openpyxl.Workbook()
        book.save('doc.xlsx')
        book.close()

# Create a method that adds the numbers to the excel file, and uses the file name as the sheet name.
# Put the list of numbers down in one column using \n as a delimiter.
def add_sheet(file_name, numbers):
    book = openpyxl.load_workbook('doc.xlsx')
    sheet = book.create_sheet(file_name)
    sheet.cell(row=1, column=1).value = 'SKUs'
    for i in range(len(numbers)):
        sheet.cell(row=i+2, column=1).value = numbers[i]
    book.save('doc.xlsx')
    book.close()




def main():
    """
    This function is the main function.
    :return:
    """
    files = os.listdir()
    create_excel_file()
    for file in files:
        if file.endswith('.txt'):
            lines = read_file(file)
            numbers = find_numbers(lines)
            add_sheet(file, numbers)
            # Print the file name and the number of 10 digit numbers
            print(file, len(numbers))



if __name__ == '__main__':
    main()


